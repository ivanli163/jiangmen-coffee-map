import os
import sqlite3
import json
import subprocess
import datetime
import time
from flask import Flask, request, jsonify, send_from_directory, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
import io

app = Flask(__name__)
CORS(app)  # 允许跨域

# --- 配置 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__)) # server/
PROJECT_ROOT = os.path.dirname(BASE_DIR) # project root

DB_PATH = os.path.join(BASE_DIR, 'database.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
SHOP_IMAGES_FOLDER = os.path.join(UPLOAD_FOLDER, 'shops')
OUTPUT_FOLDER = os.path.join(PROJECT_ROOT, 'output_jiangmen')  # 切片输出目录
MAP_IMAGE_NAME = 'current_map.png' # 统一管理地图文件名

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # 16MB max upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# --- 数据库初始化 ---
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. 地图配置表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS map_config (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    
    # 2. 区域表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS regions (
            id TEXT PRIMARY KEY,
            name TEXT,
            coord_x INTEGER,
            coord_y INTEGER
        )
    ''')
    
    # 3. 商家表 (升级版)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS shops (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            region_id TEXT,
            name TEXT,
            tag TEXT,
            rating TEXT,
            img TEXT,
            desc TEXT,
            address TEXT,
            phone TEXT,
            price TEXT,
            open_hours TEXT,
            source TEXT,
            lat REAL,
            lng REAL,
            coord_x INTEGER,
            coord_y INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(region_id) REFERENCES regions(id)
        )
    ''')

    # 4. 标签表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            color TEXT DEFAULT '#ffffff',
            bg_color TEXT DEFAULT '#d4a373'
        )
    ''')
    
    # 5. 初始化默认数据 (如果为空)
    cursor.execute('SELECT count(*) FROM map_config')
    if cursor.fetchone()[0] == 0:
        print("Initializing default data...")
        # ... (keep existing map_config initialization)
        cursor.execute("INSERT INTO map_config (key, value) VALUES (?, ?)", ('imgWidth', '3584'))
        cursor.execute("INSERT INTO map_config (key, value) VALUES (?, ?)", ('imgHeight', '4800'))
        cursor.execute("INSERT INTO map_config (key, value) VALUES (?, ?)", ('maxZoom', '5'))
        
        # 默认区域
        regions = [
            ('pengjiang', '蓬江区', 1500, 2400),
            ('jianghai', '江海区', 2500, 3000),
            ('xinhui', '新会区', 3500, 1800)
        ]
        cursor.executemany("INSERT INTO regions (id, name, coord_x, coord_y) VALUES (?, ?, ?, ?)", regions)

        # 默认标签
        tags = [
            ('网红店', '#ffffff', '#ff6b6b'),
            ('环境好', '#ffffff', '#51cf66'),
            ('必打卡', '#ffffff', '#fcc419'),
            ('老字号', '#ffffff', '#339af0'),
            ('精品咖啡', '#ffffff', '#845ef7')
        ]
        cursor.executemany("INSERT INTO tags (name, color, bg_color) VALUES (?, ?, ?)", tags)
        
        # 默认商家
        shops = [
            ('pengjiang', '邑啡·精品烘焙', '必打卡', '4.9', 'https://images.unsplash.com/photo-1497935586351-b67a49e012bf', '位于蓬江老街的宝藏店铺，主打手冲瑰夏，环境复古舒适。', '蓬江区建设路12号', 1600, 2450),
            ('jianghai', '江海时光', '环境好', '4.6', 'https://images.unsplash.com/photo-1525610553991-2bede1a236e2', '坐拥一线江景，吹着晚风喝咖啡的绝佳去处。', '江海区沿江路88号', 2600, 3100),
            ('xinhui', '陈皮咖啡馆', '特色', '5.0', 'https://images.unsplash.com/photo-1514432324607-a09d9b4aefdd', '新会特色陈皮与咖啡的完美融合，口感独特回甘。', '新会区冈州大道中101号', 3600, 1900)
        ]
        cursor.executemany("INSERT INTO shops (region_id, name, tag, rating, img, desc, address, coord_x, coord_y) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", shops)
        
    # Check if tags table is empty (migration case)
    cursor.execute('SELECT count(*) FROM tags')
    if cursor.fetchone()[0] == 0:
        print("Initializing default tags...")
        tags = [
            ('网红店', '#ffffff', '#ff6b6b'),
            ('环境好', '#ffffff', '#51cf66'),
            ('必打卡', '#ffffff', '#fcc419'),
            ('老字号', '#ffffff', '#339af0'),
            ('精品咖啡', '#ffffff', '#845ef7')
        ]
        cursor.executemany("INSERT INTO tags (name, color, bg_color) VALUES (?, ?, ?)", tags)
        conn.commit()

    conn.commit()
    conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- API 路由 ---

@app.route('/')
def index():
    return render_template('map.html')

@app.route('/map')
def map_page():
    return render_template('map.html')

@app.route('/admin')
def admin_page():
    return render_template('admin.html')

@app.route('/api/config', methods=['GET'])
def get_config():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM map_config")
    rows = cursor.fetchall()
    config = {row['key']: row['value'] for row in rows}
    conn.close()
    return jsonify(config)

@app.route('/api/debug/ls', methods=['GET'])
def debug_ls():
    try:
        root_files = os.listdir(PROJECT_ROOT)
        output_files = os.listdir(OUTPUT_FOLDER) if os.path.exists(OUTPUT_FOLDER) else "Not Found"
        return jsonify({
            'project_root': PROJECT_ROOT,
            'root_files': root_files,
            'output_folder': OUTPUT_FOLDER,
            'output_files_sample': output_files[:20] if isinstance(output_files, list) else output_files
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/data', methods=['GET'])
def get_data():
    """获取所有区域和商家数据 (供前端一次性加载)"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取区域
    cursor.execute("SELECT * FROM regions")
    regions = []
    for row in cursor.fetchall():
        regions.append({
            'id': row['id'],
            'name': row['name'],
            'coords': [row['coord_y'], row['coord_x']] # 适配前端 [Y, X] 格式
        })
        
    # 获取标签
    cursor.execute("SELECT * FROM tags")
    tags = []
    for row in cursor.fetchall():
        tags.append(dict(row))
        
    # 获取商家
    cursor.execute("SELECT * FROM shops")
    shops_dict = {}
    for row in cursor.fetchall():
        rid = row['region_id']
        if rid not in shops_dict:
            shops_dict[rid] = []
        
        # 将 row 转为 dict，确保包含所有新字段
        shop_item = dict(row)
        shops_dict[rid].append(shop_item)
        
    conn.close()
    return jsonify({'regions': regions, 'shops': shops_dict, 'tags': tags})

# --- Tags API ---

@app.route('/api/tags', methods=['GET'])
def get_tags():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM tags")
    tags = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(tags)

@app.route('/api/tags', methods=['POST'])
def add_tag():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tags (name, color, bg_color) VALUES (?, ?, ?)",
                   (data['name'], data['color'], data['bg_color']))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'status': 'success'})

@app.route('/api/tags/<int:tag_id>', methods=['PUT'])
def update_tag(tag_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. 获取旧名称
    cursor.execute("SELECT name FROM tags WHERE id=?", (tag_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Tag not found'}), 404
    old_name = row['name']
    
    # 2. 更新标签
    cursor.execute("UPDATE tags SET name=?, color=?, bg_color=? WHERE id=?",
                   (data['name'], data['color'], data['bg_color'], tag_id))
    
    # 3. 级联更新商家表 (如果改名了)
    new_name = data['name']
    if old_name != new_name:
        # 需遍历所有包含 old_name 的商家进行替换
        cursor.execute("SELECT id, tag FROM shops WHERE tag LIKE ?", (f"%{old_name}%",))
        shops_to_update = cursor.fetchall()
        
        for s in shops_to_update:
            tags_list = s['tag'].split(',') if s['tag'] else []
            if old_name in tags_list:
                # 替换
                tags_list = [new_name if t == old_name else t for t in tags_list]
                new_tags_str = ",".join(tags_list)
                cursor.execute("UPDATE shops SET tag=? WHERE id=?", (new_tags_str, s['id']))
        
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/tags/<int:tag_id>', methods=['DELETE'])
def delete_tag(tag_id):
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. 获取标签名
    cursor.execute("SELECT name FROM tags WHERE id=?", (tag_id,))
    row = cursor.fetchone()
    if row:
        tag_name = row['name']
        # 2. 从引用该标签的商家的 tag 字段中移除
        cursor.execute("SELECT id, tag FROM shops WHERE tag LIKE ?", (f"%{tag_name}%",))
        shops_to_update = cursor.fetchall()
        
        for s in shops_to_update:
            tags_list = s['tag'].split(',') if s['tag'] else []
            if tag_name in tags_list:
                tags_list.remove(tag_name)
                new_tags_str = ",".join(tags_list)
                cursor.execute("UPDATE shops SET tag=? WHERE id=?", (new_tags_str, s['id']))
        
    # 3. 删除标签
    cursor.execute("DELETE FROM tags WHERE id=?", (tag_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/regions', methods=['POST'])
def update_region():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    # coords 传来的是 [y, x]
    cursor.execute("""
        INSERT INTO regions (id, name, coord_x, coord_y) 
        VALUES (?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET 
        coord_x=excluded.coord_x, coord_y=excluded.coord_y
    """, (data['id'], data['name'], data['coords'][1], data['coords'][0]))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/upload_map', methods=['POST'])
def upload_map():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file:
        # 1. 保存文件
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], MAP_IMAGE_NAME)
        file.save(filepath)
        
        # 2. 获取图片尺寸 (使用 gdalinfo)
        try:
            result = subprocess.check_output(['gdalinfo', filepath]).decode('utf-8')
            # 简单解析 Size is 3584, 4800
            import re
            match = re.search(r'Size is (\d+), (\d+)', result)
            if match:
                width, height = match.groups()
                
                # 3. 更新数据库配置
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute("UPDATE map_config SET value=? WHERE key='imgWidth'", (width,))
                cursor.execute("UPDATE map_config SET value=? WHERE key='imgHeight'", (height,))
                conn.commit()
                conn.close()
                
                # 4. 执行切片 (后台任务)
                # 注意：为了简化，这里暂时同步执行，可能会卡顿几秒
                cmd = f"gdal2tiles.py -p raster -z 0-5 --xyz --webviewer=leaflet '{filepath}' '{OUTPUT_FOLDER}'"
                subprocess.run(cmd, shell=True, check=True)
                
                return jsonify({'status': 'success', 'width': width, 'height': height})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
    return jsonify({'error': 'Unknown error'}), 500

# --- 商家管理 API ---

@app.route('/api/shops', methods=['GET'])
def get_shops():
    region_id = request.args.get('region_id')
    conn = get_db()
    cursor = conn.cursor()
    if region_id:
        cursor.execute("SELECT * FROM shops WHERE region_id = ?", (region_id,))
    else:
        cursor.execute("SELECT * FROM shops")
    
    shops = []
    for row in cursor.fetchall():
        shops.append(dict(row))
    conn.close()
    return jsonify(shops)

@app.route('/api/shops', methods=['POST'])
def create_shop():
    data = request.json
    
    # Tag Validation: limit to 3
    tag_val = data.get('tag', '')
    if tag_val:
        tags_list = tag_val.split(',')
        if len(tags_list) > 3:
            return jsonify({'error': '最多只能选择3个标签'}), 400
        # 重新组合以清理空格等潜在问题 (可选)
        tag_val = ",".join([t.strip() for t in tags_list[:3]])
            
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO shops (region_id, name, tag, rating, img, desc, address, coord_x, coord_y, lat, lng, phone, price, open_hours, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data['region_id'], data['name'], tag_val, data['rating'], data['img'], data['desc'], data.get('address'), data.get('coord_x'), data.get('coord_y'), data.get('lat'), data.get('lng'), data.get('phone'), data.get('price'), data.get('open_hours'), data.get('source')))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'status': 'success'})

@app.route('/api/shops/<int:shop_id>', methods=['PUT'])
def update_shop(shop_id):
    data = request.json
    
    # Tag Validation: limit to 3
    tag_val = data.get('tag', '')
    if tag_val:
        tags_list = tag_val.split(',')
        if len(tags_list) > 3:
            return jsonify({'error': '最多只能选择3个标签'}), 400
        tag_val = ",".join([t.strip() for t in tags_list[:3]])
        
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE shops 
        SET region_id=?, name=?, tag=?, rating=?, img=?, desc=?, address=?, coord_x=?, coord_y=?, lat=?, lng=?, phone=?, price=?, open_hours=?, source=?
        WHERE id=?
    """, (data['region_id'], data['name'], tag_val, data['rating'], data['img'], data['desc'], data.get('address'), data.get('coord_x'), data.get('coord_y'), data.get('lat'), data.get('lng'), data.get('phone'), data.get('price'), data.get('open_hours'), data.get('source'), shop_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/shops/<int:shop_id>', methods=['DELETE'])
def delete_shop(shop_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM shops WHERE id=?", (shop_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/upload/image', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{int(time.time())}_{file.filename}")
        if not os.path.exists(SHOP_IMAGES_FOLDER):
            os.makedirs(SHOP_IMAGES_FOLDER)
        filepath = os.path.join(SHOP_IMAGES_FOLDER, filename)
        file.save(filepath)
        # 返回相对路径 URL
        url = f"/server/uploads/shops/{filename}"
        return jsonify({'url': url})
        
    return jsonify({'error': 'Invalid file'}), 400

# --- 导入导出 API ---

@app.route('/api/export/shops', methods=['GET'])
def export_shops():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM shops")
    rows = cursor.fetchall()
    
    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Shops"
    
    # 表头
    headers = ['id', 'region_id', 'name', 'tag', 'rating', 'img', 'desc', 'address', 'phone', 'price', 'open_hours', 'source', 'lat', 'lng', 'coord_x', 'coord_y']
    ws.append(headers)
    
    # 数据
    for row in rows:
        ws.append([row[h] for h in headers])
        
    conn.close()
    
    # 写入流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"shops_export_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    from flask import Response
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/api/import/shops', methods=['POST'])
def import_shops():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files are supported'}), 400
        
    try:
        # 读取 Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        required_headers = ['id', 'region_id', 'name'] # 至少需要这些
        
        # 简单校验表头
        for h in required_headers:
            if h not in headers:
                return jsonify({'error': f'Missing required header: {h}'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        stats = {'updated': 0, 'created': 0, 'failed': 0, 'errors': []}
        
        # 遍历数据行 (从第2行开始)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = dict(zip(headers, row))
                
                # 过滤 None 值
                # row_data = {k: v for k, v in row_data.items() if v is not None} 
                # 不，None 应该被处理为空字符串或保留原值? 
                # 策略: 如果 Excel 单元格为空，视为 None。数据库写入时 None 会变成 NULL。
                # 但对于字符串字段，最好转为空字符串 "" 避免 NULL 问题 (如果前端没处理的话)。
                # 这里保持简单，由 SQLite 处理。
                
                shop_id = row_data.get('id')
                
                # 检查是否存在
                exists = False
                if shop_id:
                    cursor.execute("SELECT 1 FROM shops WHERE id=?", (shop_id,))
                    if cursor.fetchone():
                        exists = True
                
                # 构建字段
                # 允许更新的字段
                update_fields = ['region_id', 'name', 'tag', 'rating', 'img', 'desc', 'address', 'phone', 'price', 'open_hours', 'source', 'lat', 'lng', 'coord_x', 'coord_y']
                
                # 准备数据字典
                data_to_save = {}
                for field in update_fields:
                    if field in row_data:
                         data_to_save[field] = row_data[field]
                
                if exists:
                    # UPDATE
                    set_clause = ", ".join([f"{k}=?" for k in data_to_save.keys()])
                    values = list(data_to_save.values())
                    values.append(shop_id)
                    
                    sql = f"UPDATE shops SET {set_clause} WHERE id=?"
                    cursor.execute(sql, values)
                    stats['updated'] += 1
                else:
                    # INSERT (如果提供了 ID，且 ID 不存在，是否允许指定 ID 插入? SQLite 自增主键通常允许手动插入)
                    # 如果没有提供 ID (None), 则自增。
                    
                    # 如果 row_data['id'] 是 None，我们就不插入 id 字段，让它自增
                    # 如果 row_data['id'] 有值，我们尝试插入它
                    
                    cols = list(data_to_save.keys())
                    vals = list(data_to_save.values())
                    
                    if shop_id:
                        cols.insert(0, 'id')
                        vals.insert(0, shop_id)
                        
                    placeholders = ", ".join(["?" for _ in cols])
                    col_names = ", ".join(cols)
                    
                    sql = f"INSERT INTO shops ({col_names}) VALUES ({placeholders})"
                    cursor.execute(sql, vals)
                    stats['created'] += 1
                    
            except Exception as e:
                stats['failed'] += 1
                stats['errors'].append(f"Row {row_idx}: {str(e)}")
                
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'stats': stats})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- 静态文件服务 ---
@app.route('/output_jiangmen/<path:filename>')
def serve_tiles(filename):
    response = send_from_directory(OUTPUT_FOLDER, filename)
    # 对于 HTML 文件，禁用缓存
    if filename.endswith('.html'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    else:
        # 其他文件 (如瓦片) 缓存 1 年
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    return response

@app.route('/server/uploads/<path:filename>')
def serve_uploads(filename):
    # 兼容处理：如果本地文件不存在，返回 404，前端会处理成 CDN 链接
    # 但由于 send_from_directory 会抛出 NotFound，我们可以捕获它
    try:
        response = send_from_directory(UPLOAD_FOLDER, filename)
        response.headers['Cache-Control'] = 'public, max-age=31536000'
        return response
    except Exception:
        return "File not found", 404

# --- 初始化 (确保在 Gunicorn 启动时也运行) ---
# 确保必要的目录存在
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(SHOP_IMAGES_FOLDER):
    os.makedirs(SHOP_IMAGES_FOLDER)
if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)

# 初始化数据库
try:
    init_db()
    print("Database initialized successfully.")
except Exception as e:
    print(f"Error initializing database: {e}")

if __name__ == '__main__':
    # 监听 0.0.0.0 允许外部访问
    app.run(host='0.0.0.0', port=8000, debug=True)
