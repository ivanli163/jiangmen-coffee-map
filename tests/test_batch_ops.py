import unittest
import requests
import os
import time

# 假设服务运行在本地 8000 端口
BASE_URL = "http://127.0.0.1:8000"

class TestCoffeeMapAdmin(unittest.TestCase):
    
    def test_1_export_shops(self):
        """测试导出 Excel"""
        print("\nTesting Export Shops...")
        response = requests.get(f"{BASE_URL}/api/export/shops")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.headers['Content-Type'].startswith('application/vnd.openxmlformats-officedocument'))
        
        # 保存文件以供查看（可选）
        with open("test_export.xlsx", "wb") as f:
            f.write(response.content)
        print("Export successful, saved to test_export.xlsx")

    def test_2_import_shops(self):
        """测试导入 Excel"""
        print("\nTesting Import Shops...")
        
        # 1. 先导出作为模板
        response = requests.get(f"{BASE_URL}/api/export/shops")
        self.assertEqual(response.status_code, 200)
        
        # 2. 修改一些数据用于测试 (使用 openpyxl)
        import openpyxl
        import io
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # 修改第一行数据 (假设至少有一行数据)
        if ws.max_row > 1:
            # 找到 'name' 列的索引
            header = [cell.value for cell in ws[1]]
            name_idx = header.index('name')
            
            # 修改第2行的 name
            original_name = ws.cell(row=2, column=name_idx+1).value
            new_name = f"{original_name} [TestUpdate]"
            ws.cell(row=2, column=name_idx+1).value = new_name
            print(f"Modifying shop: {original_name} -> {new_name}")
            
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 3. 上传导入
        files = {'file': ('test_import.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        import_res = requests.post(f"{BASE_URL}/api/import/shops", files=files)
        
        print("Import Response:", import_res.text)
        self.assertEqual(import_res.status_code, 200)
        data = import_res.json()
        self.assertEqual(data['status'], 'success')
        self.assertGreater(data['stats']['updated'], 0)
        
    def test_3_search_logic(self):
        """前端搜索逻辑测试 (模拟)"""
        # 由于搜索是前端纯 JS 实现，这里主要验证后端数据结构是否包含搜索所需字段
        print("\nTesting Data Structure for Search...")
        res = requests.get(f"{BASE_URL}/api/data")
        data = res.json()
        shops = data['shops']
        
        # 检查任意一个商家是否包含 searchable fields
        found = False
        for region_id, shop_list in shops.items():
            if shop_list:
                shop = shop_list[0]
                self.assertIn('name', shop)
                self.assertIn('address', shop) # 可选，但最好有
                # self.assertIn('tag', shop)
                found = True
                break
        
        if not found:
            print("Warning: No shops found to test structure")
        else:
            print("Data structure valid for search")

if __name__ == '__main__':
    unittest.main()
